"""Excel export service"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from datetime import datetime, date
from typing import List, Optional
from io import BytesIO

from app.models.work_session import WorkSession, WorkStatus
from app.models.user import User
from app.models.department import Department
from app.models.work_comment import WorkComment


class ExcelService:
    """Service for generating Excel reports"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def generate_department_report(
        self,
        date_from: date,
        date_to: date,
        department_ids: Optional[List[int]] = None,
        late_only: bool = False,
        include_comments: bool = True
    ) -> BytesIO:
        """Generate department report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт по подразделениям"
        
        # Header
        self._add_header(ws, "Отчёт по подразделениям", date_from, date_to)
        
        # Table headers
        headers = [
            "Подразделение", "Сотрудник", "Дата", "Начало", "Конец",
            "Работа (ч)", "Перерывы (ч)", "Опоздание (мин)"
        ]
        if include_comments:
            headers.append("Комментарий РОП")
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        query = self.db.query(WorkSession).join(User).join(Department).filter(
            WorkSession.start_time >= datetime.combine(date_from, datetime.min.time()),
            WorkSession.start_time <= datetime.combine(date_to, datetime.max.time())
        )
        
        if department_ids:
            query = query.filter(User.department_id.in_(department_ids))
        
        if late_only:
            query = query.filter(WorkSession.is_late == True)
        
        sessions = query.order_by(Department.name, User.name, WorkSession.start_time).all()
        
        row = 5
        for session in sessions:
            ws.cell(row=row, column=1, value=session.user.department.name if session.user.department else "N/A")
            ws.cell(row=row, column=2, value=session.user.name)
            ws.cell(row=row, column=3, value=session.start_time.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=session.start_time.strftime("%H:%M"))
            ws.cell(row=row, column=5, value=session.end_time.strftime("%H:%M") if session.end_time else "-")
            ws.cell(row=row, column=6, value=round(session.total_work_time / 3600, 2))
            ws.cell(row=row, column=7, value=round(session.total_break_time / 3600, 2))
            ws.cell(row=row, column=8, value=session.late_minutes if session.is_late else 0)
            
            if include_comments:
                comment = self.db.query(WorkComment).filter(
                    WorkComment.session_id == session.id
                ).first()
                ws.cell(row=row, column=9, value=comment.comment if comment else "")
            
            row += 1
        
        # Totals
        total_sessions = len(sessions)
        total_work = sum(s.total_work_time for s in sessions) / 3600
        total_breaks = sum(s.total_break_time for s in sessions) / 3600
        total_late = sum(s.late_minutes for s in sessions if s.is_late)
        
        row += 1
        ws.cell(row=row, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{total_sessions} сессий")
        ws.cell(row=row, column=6, value=round(total_work, 2))
        ws.cell(row=row, column=7, value=round(total_breaks, 2))
        ws.cell(row=row, column=8, value=total_late)
        
        # Auto-width
        self._auto_width(ws)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_employee_report(
        self,
        user_id: int,
        date_from: date,
        date_to: date,
        include_comments: bool = True
    ) -> BytesIO:
        """Generate employee report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт по сотруднику"
        
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user:
            raise ValueError("User not found")
        
        # Header
        ws.cell(row=1, column=1, value=f"Отчёт: {user.name}").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Период: {date_from} - {date_to}")
        
        # Table headers
        headers = [
            "Дата", "День", "Начало", "Конец", "Работа", "Перерывы",
            "Опоздание", "Принуд. завершение"
        ]
        if include_comments:
            headers.append("Комментарий")
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        sessions = self.db.query(WorkSession).filter(
            WorkSession.user_id == user.amocrm_user_id,
            WorkSession.start_time >= datetime.combine(date_from, datetime.min.time()),
            WorkSession.start_time <= datetime.combine(date_to, datetime.max.time())
        ).order_by(WorkSession.start_time).all()
        
        row = 5
        weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        
        for session in sessions:
            ws.cell(row=row, column=1, value=session.start_time.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=weekdays[session.start_time.weekday()])
            ws.cell(row=row, column=3, value=session.start_time.strftime("%H:%M"))
            ws.cell(row=row, column=4, value=session.end_time.strftime("%H:%M") if session.end_time else "-")
            ws.cell(row=row, column=5, value=self._format_time(session.total_work_time))
            ws.cell(row=row, column=6, value=self._format_time(session.total_break_time))
            ws.cell(row=row, column=7, value=f"{session.late_minutes} мин" if session.is_late else "-")
            ws.cell(row=row, column=8, value="Да" if session.forced_finish else "Нет")
            
            if include_comments:
                comment = self.db.query(WorkComment).filter(
                    WorkComment.session_id == session.id
                ).first()
                ws.cell(row=row, column=9, value=comment.comment if comment else "")
            
            row += 1
        
        # Totals
        total_work = sum(s.total_work_time for s in sessions)
        total_breaks = sum(s.total_break_time for s in sessions)
        late_count = sum(1 for s in sessions if s.is_late)
        
        row += 1
        ws.cell(row=row, column=1, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{len(sessions)} дней")
        ws.cell(row=row, column=5, value=self._format_time(total_work))
        ws.cell(row=row, column=6, value=self._format_time(total_breaks))
        ws.cell(row=row, column=7, value=f"{late_count} опозданий")
        
        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_late_arrivals_report(
        self,
        date_from: date,
        date_to: date,
        department_ids: Optional[List[int]] = None
    ) -> BytesIO:
        """Generate late arrivals report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Опоздания"
        
        self._add_header(ws, "Отчёт по опозданиям", date_from, date_to)
        
        headers = [
            "Дата", "Сотрудник", "Подразделение", "Начало работы",
            "Опоздание (мин)", "Причина", "Комментарий РОП"
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        query = self.db.query(WorkSession).join(User).join(Department).filter(
            WorkSession.is_late == True,
            WorkSession.start_time >= datetime.combine(date_from, datetime.min.time()),
            WorkSession.start_time <= datetime.combine(date_to, datetime.max.time())
        )
        
        if department_ids:
            query = query.filter(User.department_id.in_(department_ids))
        
        sessions = query.order_by(WorkSession.start_time.desc()).all()
        
        row = 5
        for session in sessions:
            ws.cell(row=row, column=1, value=session.start_time.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=session.user.name)
            ws.cell(row=row, column=3, value=session.user.department.name if session.user.department else "N/A")
            ws.cell(row=row, column=4, value=session.start_time.strftime("%H:%M"))
            ws.cell(row=row, column=5, value=session.late_minutes)
            ws.cell(row=row, column=6, value=session.late_reason or "-")
            
            comment = self.db.query(WorkComment).filter(
                WorkComment.session_id == session.id
            ).first()
            ws.cell(row=row, column=7, value=comment.comment if comment else "")
            
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="ИТОГО опозданий:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(sessions))
        ws.cell(row=row, column=5, value=sum(s.late_minutes for s in sessions))
        
        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _add_header(self, ws, title: str, date_from: date, date_to: date):
        """Add report header"""
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Период: {date_from} - {date_to}")
        ws.cell(row=3, column=1, value=f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    def _format_time(self, seconds: int) -> str:
        """Format seconds to HH:MM"""
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        return f"{hours}:{minutes:02d}"
    
    def _auto_width(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
